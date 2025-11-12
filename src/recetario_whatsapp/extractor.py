"""
Extractor de recetas desde archivos de chat de WhatsApp.
"""

import re
import json
import os
import argparse
import unicodedata
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple, Set

from openpyxl import load_workbook
from .mistral_client import MistralClient
from .supabase_utils import SupabaseManager

# Importar pandas y openpyxl para procesamiento de Excel
try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warning: pandas not available. Excel processing will not work.")


class ExcelExtractor:
    """Extractor de recetas desde archivos Excel."""

    def __init__(self, supabase_manager: SupabaseManager):
        """Inicializa el extractor de Excel."""
        self.supabase_manager = supabase_manager
        self.existing_keys: Set[Tuple[str, str]] = set()
        self.nuevas_claves: Set[Tuple[str, str]] = set()
        self.creador_aliases = {
            # Alias comunes para evitar duplicados por variaciones de nombre
            "carlitos": "carlos",
        }
        self._refrescar_claves_existentes()

    def procesar_excel(self, ruta_archivo: str) -> Dict[str, Any]:
        """
        Procesa un archivo Excel y extrae recetas.

        Args:
            ruta_archivo: Ruta al archivo Excel

        Returns:
            Diccionario con estadísticas del procesamiento
        """
        if not PANDAS_AVAILABLE:
            return {
                "error": "pandas not available. Install with: pip install pandas openpyxl"
            }

        print(f"Procesando archivo Excel: {ruta_archivo}")

        # Refrescar deduplicación y limpiar cache por ejecución
        self._refrescar_claves_existentes()
        self.nuevas_claves.clear()

        try:
            # Leer todas las hojas del Excel (datos tabulares)
            excel_data = pd.read_excel(ruta_archivo, sheet_name=None)
            # Cargar workbook con openpyxl para inspeccionar imágenes embebidas
            workbook = load_workbook(ruta_archivo, data_only=True)
            print(f"Encontradas {len(excel_data)} hojas")

            recetas_extraidas = 0
            recetas_insertadas = 0
            hojas_procesadas = 0

            # Procesar cada hoja
            for sheet_name, sheet_data in excel_data.items():
                print(f"Procesando hoja: {sheet_name}")
                hojas_procesadas += 1

                # Extraer recetas de esta hoja
                imagenes_por_fila, imagenes_sin_posicion = (
                    self._obtener_imagenes_embebidas(workbook, sheet_name)
                )
                resultado_hoja = self._extraer_recetas_de_hoja(
                    sheet_data,
                    sheet_name,
                    imagenes_por_fila,
                    imagenes_sin_posicion,
                )

                recetas_extraidas += resultado_hoja["recetas_extraidas"]
                recetas_insertadas += resultado_hoja["recetas_insertadas"]

                print(
                    f"  Hoja '{sheet_name}': {resultado_hoja['recetas_extraidas']} recetas extraídas, {resultado_hoja['recetas_insertadas']} insertadas"
                )

            return {
                "hojas_procesadas": hojas_procesadas,
                "recetas_extraidas": recetas_extraidas,
                "recetas_insertadas": recetas_insertadas,
                "archivo_tipo": "excel",
            }

        except Exception as e:
            print(f"Error procesando Excel: {e}")
            return {"error": str(e)}

    def _extraer_recetas_de_hoja(
        self,
        hoja_data: pd.DataFrame,
        hoja_name: str,
        imagenes_por_fila: Dict[int, List[Dict[str, Any]]],
        imagenes_sin_posicion: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """
        Extrae recetas de una hoja específica del Excel.

        Args:
            hoja_data: DataFrame con los datos de la hoja
            hoja_name: Nombre de la hoja

        Returns:
            Diccionario con estadísticas de la hoja
        """
        recetas_extraidas = 0
        recetas_insertadas = 0

        hoja_data = hoja_data.fillna("")
        columnas_originales = [
            str(col) if col is not None else "" for col in hoja_data.columns
        ]
        hoja_data.columns = [col.strip() for col in columnas_originales]

        # Si no hay filas pero sí columnas con datos, crear una fila sintética a partir de los encabezados
        if hoja_data.empty and any(col.strip() for col in hoja_data.columns):
            fila_sintetica = {col: col for col in hoja_data.columns}
            hoja_data = pd.DataFrame([fila_sintetica])

        columnas = list(hoja_data.columns)

        # Buscar recetas en la hoja
        # El nombre de la receta debe estar en la primera columna
        # Los ingredientes y preparación en las columnas siguientes

        for idx, (row_idx, row) in enumerate(hoja_data.iterrows()):
            # Obtener el nombre de la receta de la primera columna
            nombre_receta = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""

            # Saltar filas vacías o que empiecen con texto de encabezado
            if not nombre_receta or self._es_encabezado(nombre_receta):
                continue

            # Verificar que no sea un link o imagen (según la solicitud del usuario)
            if self._es_link_o_imagen(row):
                print(f"  ⚠️ Saltando fila {idx+1}: parece ser un link o imagen")
                continue

            creador = self._obtener_creador(row, columnas, hoja_name)
            clave_normalizada = (
                self._normalizar_texto(creador),
                self._normalizar_texto(nombre_receta),
            )

            if (
                clave_normalizada in self.existing_keys
                or clave_normalizada in self.nuevas_claves
            ):
                print(
                    f"  ⚠️ Receta duplicada detectada: '{nombre_receta}' de {creador}. Saltando."
                )
                continue

            # Extraer ingredientes y preparación de las columnas siguientes
            ingredientes = self._extraer_ingredientes(row, columnas)
            preparacion = self._extraer_preparacion(row, columnas)
            imagenes = self._extraer_imagenes_por_valor(row, columnas, creador)

            # Adjuntar imágenes embebidas en el Excel si existen
            imagenes_receta = imagenes.copy()
            imagenes_embebidas = self._resolver_imagenes_embebidas(
                imagenes_por_fila,
                imagenes_sin_posicion,
                row_idx,
                nombre_receta,
                creador,
            )
            if imagenes_embebidas:
                imagenes_receta.extend(imagenes_embebidas)

            # Si no hay ingredientes, saltar esta fila
            if not ingredientes:
                continue

            # Crear receta
            receta = {
                "creador": creador,
                "nombre_receta": nombre_receta,
                "ingredientes": ingredientes,
                "pasos_preparacion": preparacion if preparacion else None,
                "tiene_foto": bool(imagenes_receta),
                "fecha_mensaje": datetime.now().isoformat(),
                "imagenes": imagenes_receta,
                "url_imagen": imagenes_receta[0]["url"] if imagenes_receta else None,
            }

            # Insertar en la base de datos
            if self.supabase_manager.insertar_receta(receta):
                recetas_insertadas += 1
                self.nuevas_claves.add(clave_normalizada)
                self.existing_keys.add(clave_normalizada)
                print(f"  ✅ Receta '{nombre_receta}' insertada")
            else:
                print(f"  ❌ Error insertando receta '{nombre_receta}'")

            recetas_extraidas += 1

        return {
            "recetas_extraidas": recetas_extraidas,
            "recetas_insertadas": recetas_insertadas,
        }

    def _es_encabezado(self, texto: str) -> bool:
        """Determina si el texto parece ser un encabezado."""
        texto_lower = texto.lower()
        encabezados = [
            "ingredientes",
            "preparación",
            "pasos",
            "método",
            "receta",
            "preparacion",
            "metodología",
            "instrucciones",
            "elaboración",
        ]
        return any(encabezado in texto_lower for encabezado in encabezados)

    def _es_link_o_imagen(self, row: pd.Series) -> bool:
        """Determina si la fila completa solo contiene un link o imagen (sin título)."""
        valores_texto = [
            str(value).strip()
            for value in row
            if pd.notna(value) and str(value).strip()
        ]
        if not valores_texto:
            return False

        # Si solo hay una celda y parece URL/imagen, considerarlo fila vacía
        if len(valores_texto) == 1:
            return self._es_url(valores_texto[0])

        # Si hay múltiples celdas, no bloquear toda la fila; la detección se hará por célula
        return False

    def _obtener_creador(
        self, row: pd.Series, columns: List[str], hoja_name: str
    ) -> str:
        """Determina el creador de la receta."""
        for col_name in columns:
            col_lower = col_name.lower()
            if any(palabra in col_lower for palabra in ["autor", "creador"]):
                value = row[col_name]
                if pd.notna(value) and str(value).strip():
                    return self._aplicar_alias_creador(str(value).strip())

        # Si no hay columna específica, usar el nombre de la hoja como creador
        if hoja_name:
            return self._aplicar_alias_creador(hoja_name.strip())

        return "Excel Import"

    def _extraer_ingredientes(self, row: pd.Series, columns: List[str]) -> str:
        """Extrae los ingredientes de una fila."""
        ingredientes = []

        # Buscar en columnas que contengan "ingredientes" en el nombre
        for col_name in columns:
            if "ingredientes" in col_name.lower():
                value = row[col_name]
                if pd.notna(value) and str(value).strip():
                    ingredientes.append(str(value).strip())

        # Si no encontró columna específica, usar la segunda columna
        if not ingredientes and len(columns) > 1:
            value = row.iloc[1]
            if pd.notna(value) and str(value).strip():
                ingredientes.append(str(value).strip())

        # Si aún no hay ingredientes pero hay más columnas, buscar en la tercera
        if not ingredientes and len(columns) > 2:
            value = row.iloc[2]
            if pd.notna(value) and str(value).strip() and not self._es_url(str(value)):
                ingredientes.append(str(value).strip())

        return "\n".join(ingredientes) if ingredientes else ""

    def _extraer_preparacion(self, row: pd.Series, columns: List[str]) -> str:
        """Extrae la preparación de una fila."""
        preparacion = []

        # Buscar en columnas que contengan palabras de preparación
        for col_name in columns:
            col_lower = col_name.lower()
            if any(
                palabra in col_lower
                for palabra in [
                    "preparación",
                    "preparacion",
                    "pasos",
                    "método",
                    "metodología",
                    "elaboración",
                    "instrucciones",
                ]
            ):
                value = row[col_name]
                if pd.notna(value) and str(value).strip():
                    preparacion.append(str(value).strip())

        # Si no encontró columna específica, buscar en columnas después de la segunda
        if not preparacion:
            for i in range(2, min(5, len(columns))):  # Buscar hasta la quinta columna
                value = row.iloc[i]
                if pd.notna(value) and str(value).strip():
                    # Solo agregar si parece ser pasos de preparación (no ingredientes)
                    str_value = str(value).strip()
                    if not self._parece_ingrediente(str_value) and not self._es_url(
                        str_value
                    ):
                        preparacion.append(str_value)

        return "\n".join(preparacion) if preparacion else ""

    def _extraer_imagenes_por_valor(
        self, row: pd.Series, columns: List[str], creador: str
    ) -> List[Dict[str, Any]]:
        """Detecta URLs de imagen en la fila (por ejemplo, tercera columna)."""
        imagenes: List[Dict[str, Any]] = []

        for col_name in columns:
            col_lower = col_name.lower()
            if (
                any(palabra in col_lower for palabra in ["imagen", "foto", "url"])
                or "imagenes" in col_lower
            ):
                value = row[col_name]
                if pd.notna(value):
                    posibles_urls = self._extraer_urls(str(value))
                    imagenes.extend(
                        self._construir_objetos_imagen(posibles_urls, creador)
                    )

        # Si no hay columnas específicas, revisar las columnas siguientes a ingredientes/preparación
        for i in range(2, len(columns)):
            value = row.iloc[i]
            if pd.notna(value):
                posibles_urls = self._extraer_urls(str(value))
                imagenes.extend(self._construir_objetos_imagen(posibles_urls, creador))

        # Eliminar duplicados manteniendo orden
        vistas = set()
        imagenes_unicas = []
        for img in imagenes:
            url = img["url"]
            if url not in vistas:
                vistas.add(url)
                imagenes_unicas.append(img)
        return imagenes_unicas

    def _extraer_urls(self, texto: str) -> List[str]:
        if not texto:
            return []
        patrones = re.findall(r"(https?://\S+)", texto, re.IGNORECASE)
        # Limpiar signos de puntuación al final
        urls = [url.rstrip(".,);]\"'") for url in patrones]
        return [url for url in urls if self._es_url(url)]

    def _construir_objetos_imagen(
        self, urls: List[str], creador: str
    ) -> List[Dict[str, Any]]:
        imagenes = []
        for url in urls:
            imagenes.append(
                {
                    "url": url,
                    "autor": creador,
                    "descripcion": None,
                }
            )
        return imagenes

    @staticmethod
    def _es_url(texto: str) -> bool:
        if not texto:
            return False
        texto = texto.strip().lower()
        patrones = [
            "http://",
            "https://",
            "www.",
            ".jpg",
            ".jpeg",
            ".png",
            ".gif",
            ".webp",
        ]
        return any(patron in texto for patron in patrones)

    def _obtener_imagenes_embebidas(
        self, workbook, sheet_name: str
    ) -> Tuple[Dict[int, List[Dict[str, Any]]], List[Dict[str, Any]]]:
        imagenes_por_fila: Dict[int, List[Dict[str, Any]]] = {}
        imagenes_sin_posicion: List[Dict[str, Any]] = []

        if sheet_name not in workbook.sheetnames:
            return imagenes_por_fila, imagenes_sin_posicion

        worksheet = workbook[sheet_name]
        imagenes = getattr(worksheet, "_images", [])

        for idx, image in enumerate(imagenes, start=1):
            imagen_bytes = self._obtener_bytes_imagen(image)
            if not imagen_bytes:
                continue

            formato = getattr(image, "format", "png") or "png"
            fila = self._obtener_fila_desde_anchor(image.anchor)

            info_imagen = {
                "bytes": imagen_bytes,
                "formato": formato.lower(),
                "indice": idx,
            }

            if fila is not None:
                imagenes_por_fila.setdefault(fila, []).append(info_imagen)
            else:
                imagenes_sin_posicion.append(info_imagen)

        return imagenes_por_fila, imagenes_sin_posicion

    def _obtener_fila_desde_anchor(self, anchor: Any) -> Optional[int]:
        if not anchor:
            return None

        origen = (
            getattr(anchor, "_from", None) or getattr(anchor, "from", None) or anchor
        )

        fila = getattr(origen, "row", None)
        if fila is None:
            return None

        # openpyxl usa índices base 0; DataFrame comienza en 0 tras eliminar encabezados (fila 0 es encabezado en Excel)
        fila_df = int(fila) - 1
        return fila_df if fila_df >= 0 else None

    @staticmethod
    def _obtener_bytes_imagen(image) -> Optional[bytes]:
        data = None
        if hasattr(image, "_data"):
            try:
                data = image._data()
            except Exception:
                data = None

        if hasattr(image, "image") and hasattr(image.image, "tobytes"):
            try:
                return image.image.tobytes()
            except Exception:
                pass

        if hasattr(image, "ref") and hasattr(image.ref, "blob"):
            return image.ref.blob

        if data is None:
            return None

        if isinstance(data, bytes):
            return data
        if hasattr(data, "getvalue"):
            return data.getvalue()
        if hasattr(data, "read"):
            return data.read()
        return None

    def _resolver_imagenes_embebidas(
        self,
        imagenes_por_fila: Dict[int, List[Dict[str, Any]]],
        imagenes_sin_posicion: List[Dict[str, Any]],
        row_idx: int,
        nombre_receta: str,
        creador: str,
    ) -> List[Dict[str, Any]]:
        imagenes_resultado: List[Dict[str, Any]] = []

        imagenes_fila = imagenes_por_fila.pop(row_idx, [])

        if not imagenes_fila:
            return imagenes_resultado

        if not self.supabase_manager.imagenes_habilitadas():
            print("  ⚠️ Cloudinary no disponible: no se subirán imágenes embebidas")
            return imagenes_resultado

        for posicion, info in enumerate(imagenes_fila, start=1):
            upload = self._subir_imagen_embebida(info, nombre_receta, creador, posicion)
            if upload:
                upload["autor"] = creador
                imagenes_resultado.append(upload)

        return imagenes_resultado

    def _subir_imagen_embebida(
        self,
        info_imagen: Dict[str, Any],
        nombre_receta: str,
        creador: str,
        posicion: int,
    ) -> Optional[Dict[str, Any]]:
        bytes_imagen = info_imagen.get("bytes")
        if not bytes_imagen:
            return None

        extension = info_imagen.get("formato", "png")
        nombre_archivo = self._generar_nombre_imagen(
            nombre_receta, creador, posicion, extension
        )

        resultado = self.supabase_manager.subir_imagen(bytes_imagen, nombre_archivo)
        if resultado:
            return resultado

        print(f"  ⚠️ Error subiendo imagen embebida '{nombre_archivo}'")
        return None

    @staticmethod
    def _generar_nombre_imagen(
        nombre_receta: str, creador: str, posicion: int, extension: str
    ) -> str:
        def slugify(valor: str) -> str:
            valor = (
                unicodedata.normalize("NFKD", valor)
                .encode("ascii", "ignore")
                .decode("ascii")
            )
            valor = re.sub(r"[^a-zA-Z0-9]+", "-", valor).strip("-").lower()
            return valor or "imagen"

        base = f"{slugify(creador)}-{slugify(nombre_receta)}-{posicion}"
        return f"{base}.{extension.strip('.')}"

    def _parece_ingrediente(self, texto: str) -> bool:
        """Determina si el texto parece ser una lista de ingredientes."""
        texto_lower = texto.lower()
        # Si contiene muchas cantidades o unidades de medida, probablemente es ingredientes
        patrones_cantidad = r"\d+\s*(g|kg|ml|l|lt|cucharadas?|cuch|cdas?|tazas?|onzas?|piezas?|unidades?)"
        return len(re.findall(patrones_cantidad, texto_lower)) > 2

    def _aplicar_alias_creador(self, creador: str) -> str:
        normalizado = self._normalizar_texto(creador)
        alias_objetivo = self.creador_aliases.get(normalizado)
        if alias_objetivo:
            return alias_objetivo
        return creador

    @staticmethod
    def _normalizar_texto(texto: Optional[str]) -> str:
        if not texto:
            return ""
        texto = str(texto).strip().lower()
        texto = unicodedata.normalize("NFKD", texto)
        return "".join(c for c in texto if not unicodedata.combining(c))

    def _refrescar_claves_existentes(self) -> None:
        try:
            self.existing_keys = self.supabase_manager.obtener_claves_recetas()
        except Exception as exc:
            print(f"Error refrescando claves existentes: {exc}")
            self.existing_keys = set()


class WhatsAppExtractor:
    """Extractor de recetas desde archivos de WhatsApp."""

    def __init__(self):
        """Inicializa el extractor con los clientes necesarios."""
        self.mistral_client = MistralClient()
        self.supabase_manager = SupabaseManager()
        self.excel_extractor = (
            ExcelExtractor(self.supabase_manager) if PANDAS_AVAILABLE else None
        )

        # Patrones para detectar mensajes de WhatsApp - formato real del archivo
        self.patron_mensaje = re.compile(
            r"\[(\d{2}/\d{2}/\d{2}),\s*(\d{2}:\d{2}:\d{2})\]\s*([^:]+):\s*(.*)"
        )
        self.patron_mensaje_alternativo = re.compile(
            r"\[(\d{2}/\d{2}/\d{2}),\s*(\d{2}:\d{2})\]\s*([^:]+):\s*(.*)"
        )

        # Patrón para el formato real del archivo: DD/MM/YY, HH:MM - Nombre: mensaje
        self.patron_mensaje_real = re.compile(
            r"(\d{2}/\d{2}/\d{2}),\s*(\d{2}:\d{2})\s*-\s*([^:]+):\s*(.*)"
        )

    def procesar_archivo(
        self, ruta_archivo: str, fecha_desde: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Procesa un archivo y extrae recetas. Detecta automáticamente si es WhatsApp o Excel.

        Args:
            ruta_archivo: Ruta al archivo
            fecha_desde: Fecha desde la cual procesar (formato YYYY-MM-DD) - solo para WhatsApp

        Returns:
            Diccionario con estadísticas del procesamiento
        """
        print(f"Procesando archivo: {ruta_archivo}")

        # Detectar tipo de archivo por extensión
        if ruta_archivo.lower().endswith((".xlsx", ".xls")):
            # Procesar como Excel
            if not self.excel_extractor:
                return {
                    "error": "Excel support not available. Install pandas and openpyxl."
                }
            return self.excel_extractor.procesar_excel(ruta_archivo)
        else:
            # Procesar como WhatsApp
            return self._procesar_whatsapp(ruta_archivo, fecha_desde)

    def _procesar_whatsapp(
        self, ruta_archivo: str, fecha_desde: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Procesa un archivo de WhatsApp y extrae recetas.

        Args:
            ruta_archivo: Ruta al archivo de WhatsApp
            fecha_desde: Fecha desde la cual procesar (formato YYYY-MM-DD)

        Returns:
            Diccionario con estadísticas del procesamiento
        """
        print(f"Procesando archivo: {ruta_archivo}")

        # Leer archivo
        try:
            with open(ruta_archivo, "r", encoding="utf-8") as f:
                contenido = f.read()
        except Exception as e:
            print(f"Error leyendo archivo: {e}")
            return {"error": str(e)}

        # Parsear mensajes
        mensajes = self._parsear_mensajes(contenido)
        print(f"Encontrados {len(mensajes)} mensajes")

        # Determinar fecha mínima a procesar
        fecha_limite = fecha_desde
        if not fecha_limite:
            fecha_limite = self._obtener_fecha_guardada()
            if fecha_limite:
                print(f"Usando fecha guardada: {fecha_limite}")

        # Filtrar por fecha si se especifica
        if fecha_limite:
            mensajes = self._filtrar_por_fecha(mensajes, fecha_limite)
            print(f"Después del filtro desde {fecha_limite}: {len(mensajes)} mensajes")

        # Agrupar mensajes consecutivos
        bloques = self._agrupar_mensajes_consecutivos(mensajes)
        print(f"Agrupados en {len(bloques)} bloques")

        # Procesar cada bloque
        recetas_extraidas = 0
        recetas_insertadas = 0

        for bloque in bloques:
            print(f"Procesando bloque grande ({len(bloque['texto'])} caracteres)")

            # Mostrar tokens aproximados
            tokens_aprox = len(bloque["texto"]) // 4
            print(f"  🔢 Tokens aproximados: {tokens_aprox}")

            # Extraer recetas con Mistral (ahora devuelve múltiples recetas)
            resultado = self.mistral_client.extraer_receta(bloque["texto"])

            if resultado.get("error"):
                print(f"  ❌ Error procesando bloque: {resultado['error']}")
                continue

            # Procesar todas las recetas encontradas en el bloque
            recetas_en_bloque = resultado.get("recetas", [])

            if recetas_en_bloque:
                print(f"  Encontradas {len(recetas_en_bloque)} recetas en el bloque")

                for receta in recetas_en_bloque:
                    # El nuevo formato ya viene con recetas válidas directamente
                    recetas_extraidas += 1
                    print(
                        f"  ✅ Receta: {receta.get('nombre_receta', 'Sin nombre')} de {receta.get('creador')}"
                    )

                    # Preparar datos para Supabase
                    datos_receta = {
                        "creador": receta.get("creador"),
                        "nombre_receta": receta.get("nombre_receta"),
                        "ingredientes": receta.get("ingredientes"),
                        "pasos_preparacion": receta.get("pasos_preparacion"),
                        "tiene_foto": receta.get("tiene_foto", False),
                        "url_imagen": None,
                        "fecha_mensaje": receta.get("fecha_mensaje"),
                    }

                    # Insertar en Supabase
                    if self.supabase_manager.insertar_receta(datos_receta):
                        recetas_insertadas += 1
                    else:
                        print(f"  ❌ Error insertando receta")
            else:
                print(f"  ℹ️ No se encontraron recetas en el bloque")

        # Actualizar estado de procesamiento
        self._actualizar_estado_procesamiento(
            mensajes[-1]["fecha"] if mensajes else None
        )

        return {
            "mensajes_procesados": len(mensajes),
            "bloques_procesados": len(bloques),
            "recetas_extraidas": recetas_extraidas,
            "recetas_insertadas": recetas_insertadas,
        }

    def _parsear_mensajes(self, contenido: str) -> List[Dict[str, Any]]:
        """Parsea los mensajes del archivo de WhatsApp."""
        mensajes = []
        lineas = contenido.split("\n")

        for linea in lineas:
            linea = linea.strip()
            if not linea:
                continue

            # Intentar patrón principal
            match = self.patron_mensaje.match(linea)
            if match:
                fecha_str, hora_str, creador, mensaje = match.groups()
                fecha_completa = f"{fecha_str} {hora_str}"
                mensajes.append(
                    {
                        "fecha": fecha_completa,
                        "creador": creador.strip(),
                        "mensaje": mensaje.strip(),
                    }
                )
                continue

            # Intentar patrón alternativo
            match = self.patron_mensaje_alternativo.match(linea)
            if match:
                fecha_str, hora_str, creador, mensaje = match.groups()
                fecha_completa = f"{fecha_str} {hora_str}:00"
                mensajes.append(
                    {
                        "fecha": fecha_completa,
                        "creador": creador.strip(),
                        "mensaje": mensaje.strip(),
                    }
                )
                continue

            # Intentar patrón real del archivo: DD/MM/YY, HH:MM - Nombre: mensaje
            match = self.patron_mensaje_real.match(linea)
            if match:
                fecha_str, hora_str, creador, mensaje = match.groups()
                fecha_completa = f"{fecha_str} {hora_str}:00"
                mensajes.append(
                    {
                        "fecha": fecha_completa,
                        "creador": creador.strip(),
                        "mensaje": mensaje.strip(),
                    }
                )
                continue

            # Si no coincide con ningún patrón, podría ser una línea de receta sin formato
            # Buscar si la línea anterior fue una receta y esta es continuación
            if mensajes and self._es_linea_receta(linea):
                # Agregar como continuación de la última receta
                ultimo_mensaje = mensajes[-1]
                ultimo_mensaje["mensaje"] += f"\n{linea.strip()}"
                continue

        return mensajes

    def _filtrar_por_fecha(
        self, mensajes: List[Dict[str, Any]], fecha_desde: str
    ) -> List[Dict[str, Any]]:
        """Filtra mensajes desde una fecha específica."""
        try:
            fecha_limite = datetime.strptime(fecha_desde, "%Y-%m-%d")
            mensajes_filtrados = []

            for mensaje in mensajes:
                try:
                    # Convertir fecha del mensaje (formato DD/MM/YY)
                    fecha_mensaje = datetime.strptime(
                        mensaje["fecha"].split()[0], "%d/%m/%y"
                    )
                    if fecha_mensaje >= fecha_limite:
                        mensajes_filtrados.append(mensaje)
                except ValueError:
                    # Si no se puede parsear la fecha, incluir el mensaje
                    mensajes_filtrados.append(mensaje)

            return mensajes_filtrados

        except ValueError:
            print(
                f"Error en formato de fecha: {fecha_desde}. Usando formato YYYY-MM-DD"
            )
            return mensajes

    def _agrupar_mensajes_consecutivos(
        self, mensajes: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """Agrupa mensajes en bloques más inteligentes basados en conversación y contenido."""
        if not mensajes:
            return []

        bloques = []

        # Procesar mensajes de manera más granular
        i = 0
        while i < len(mensajes):
            mensaje_actual = mensajes[i]
            texto_actual = mensaje_actual["mensaje"].lower()
            es_receta_actual = self._es_mensaje_receta(texto_actual)

            # Si no es receta, procesar siguiente
            if not es_receta_actual:
                i += 1
                continue

            # Encontramos un mensaje que podría ser receta, buscar mensajes relacionados
            bloque = {
                "creador": mensaje_actual["creador"],
                "fecha": mensaje_actual["fecha"],
                "texto": f"[{mensaje_actual['fecha']}] {mensaje_actual['creador']}: {mensaje_actual['mensaje']}\n",
            }

            # Buscar mensajes consecutivos del mismo autor que puedan ser parte de la misma receta
            j = i + 1
            while j < len(mensajes):
                siguiente_mensaje = mensajes[j]

                # Si cambió de autor o pasaron más de 5 mensajes, parar
                if (
                    siguiente_mensaje["creador"] != mensaje_actual["creador"]
                    or j - i > 5
                ):  # Máximo 5 mensajes consecutivos
                    break

                # Verificar si el siguiente mensaje parece relacionado con receta
                texto_siguiente = siguiente_mensaje["mensaje"].lower()
                if self._es_mensaje_receta_continuacion(texto_siguiente):
                    bloque[
                        "texto"
                    ] += f"[{siguiente_mensaje['fecha']}] {siguiente_mensaje['creador']}: {siguiente_mensaje['mensaje']}\n"
                    j += 1
                else:
                    break

            # Agregar bloque si tiene contenido
            if bloque["texto"].strip():
                bloques.append(bloque)

            i = j  # Saltar los mensajes que ya procesamos

        return bloques

    def _es_mensaje_receta(self, texto: str) -> bool:
        """Determina si un mensaje individual parece contener una receta."""
        texto_lower = texto.lower()

        # Palabras clave que indican inicio de receta
        indicadores_receta = [
            "receta:",
            "receta de",
            "ingredientes:",
            "receta",  # Permitir "receta" suelto ahora
        ]

        # Palabras que sugieren nombres de recetas
        palabras_receta = [
            "estofado",
            "flan",
            "torta",
            "salsa",
            "guiso",
            "sopa",
            "ensalada",
            "pasta",
            "arroz",
            "pescado",
            "carne",
            "pollo",
            "verduras",
            "costilla",
            "milagros",
            "teriyaki",
            "césar",
            "patatas",
        ]

        # Patrones de cantidades (números seguidos de unidades de medida)
        patrones_cantidad = [
            r"\d+\s*(g|kg|ml|l|lt|onzas?|onz|tazas?|taza|cucharadas?|cucharada|cuch|cdas?|cdita)",
            r"\d+\s*(piezas?|unidades?|uds?\.?|pz|pz\.?)",
            r"\d+\s*(kg|gr|mg|cl|dl)",  # Más unidades
        ]
        tiene_cantidades = any(
            re.search(patron, texto, re.IGNORECASE) for patron in patrones_cantidad
        )

        # Verificar si parece nombre de receta
        tiene_palabras_receta = any(
            palabra in texto_lower for palabra in palabras_receta
        )

        # Verificar indicadores específicos
        tiene_indicador_especifico = any(
            indicador in texto_lower for indicador in indicadores_receta
        )

        # Considerar receta si:
        # 1. Tiene indicador específico
        # 2. Tiene palabras de receta + cantidades
        # 3. Tiene palabras de receta + parece nombre corto
        if tiene_indicador_especifico:
            return True

        if tiene_palabras_receta and (tiene_cantidades or len(texto.split()) <= 3):
            return True

        # Para casos sin indicador específico, ser más permisivo
        palabras_cocina = [
            "hornear",
            "cocinar",
            "mezclar",
            "batir",
            "freír",
            "asar",
            "hervir",
            "15min",
            "olla",
        ]
        tiene_palabras_cocina = any(
            palabra in texto_lower for palabra in palabras_cocina
        )

        return tiene_cantidades and tiene_palabras_cocina

    def _es_mensaje_receta_continuacion(self, texto: str) -> bool:
        """Determina si un mensaje parece ser continuación de una receta."""
        texto_lower = texto.lower()

        # Palabras clave de continuación
        continuacion = [
            "ingredientes",
            "pasos",
            "preparación",
            "receta",
            "-",
            "•",
            "*",  # Marcadores de lista
            "mezclar",
            "hornear",
            "cocinar",
            "batir",
            "revolver",  # Acciones de cocina
        ]

        # Si tiene guiones al inicio (listas)
        tiene_guiones = bool(re.search(r"^\s*[-•*]\s", texto))

        # Si tiene números al inicio (pasos numerados)
        tiene_numeros = bool(re.search(r"^\s*\d+\.?\s*", texto))

        # Si tiene palabras de continuación
        tiene_palabras = any(palabra in texto_lower for palabra in continuacion)

        # Si tiene cantidades
        tiene_cantidades = bool(
            re.search(
                r"\d+\s*(g|kg|ml|l|lt|cucharadas?|cuch|cdas?|tazas?|onzas?|piezas?)",
                texto,
                re.IGNORECASE,
            )
        )

        return tiene_guiones or tiene_numeros or tiene_palabras or tiene_cantidades

    def _es_linea_receta(self, texto: str) -> bool:
        """Determina si una línea sin formato de WhatsApp podría ser parte de una receta."""
        texto_lower = texto.lower().strip()

        # Si está vacía, no es receta
        if not texto_lower:
            return False

        # Si parece un mensaje del sistema de WhatsApp, no es receta
        if any(
            palabra in texto_lower
            for palabra in ["multimedia omitido", "cifrado", "extremo a extremo"]
        ):
            return False

        # Si tiene cantidades, probablemente es receta
        if re.search(
            r"\d+\s*(g|kg|ml|l|lt|cucharadas?|cuch|cdas?|tazas?|onzas?|piezas?)",
            texto,
            re.IGNORECASE,
        ):
            return True

        # Si tiene guiones al inicio (listas)
        if re.search(r"^\s*[-•*]\s", texto):
            return True

        # Si tiene números al inicio (pasos numerados)
        if re.search(r"^\s*\d+\.?\s*", texto):
            return True

        # Palabras que sugieren ingredientes o pasos
        palabras_receta = [
            "hornear",
            "cocinar",
            "mezclar",
            "batir",
            "freír",
            "asar",
            "hervir",
            "min",
            "hora",
            "minutos",
            "pasos",
            "preparación",
            "olla",
            "sartén",
            "tomates",
            "cebolla",
            "ajo",
            "pimiento",
            "patatas",
            "carne",
            "pollo",
            "pescado",
            "arroz",
            "pasta",
            "salsa",
            "estofado",
            "guiso",
            "sopa",
        ]

        return any(palabra in texto_lower for palabra in palabras_receta)

    def _es_candidato_receta(self, texto: str) -> bool:
        """Determina si un bloque de texto puede contener recetas."""
        # Para bloques grandes, siempre procesamos (más eficiente que filtrar)
        # ya que Mistral puede identificar mejor qué es receta y qué no
        return True

        # Comentado el código anterior para simplificar
        """
        texto_lower = texto.lower()

        # Palabras clave que indican receta
        palabras_clave = ['ingredientes', 'receta', 'pasos', 'preparación']
        tiene_palabras_clave = any(palabra in texto_lower for palabra in palabras_clave)

        # Patrones de lista (guiones, bullets, cantidades)
        patrones_lista = [
            r'^\s*[-•]\s+',  # Guiones o bullets
            r'\d+\s*[gkgmltazas]',  # Cantidades con unidades
            r'\d+\s+[a-zA-Z]',  # Número seguido de palabra
        ]
        tiene_patrones_lista = any(re.search(patron, texto, re.MULTILINE) for patron in patrones_lista)

        return tiene_palabras_clave or tiene_patrones_lista
        """

    def _actualizar_estado_procesamiento(self, ultima_fecha: Optional[str]):
        """Actualiza el archivo de estado con la última fecha procesada."""
        estado = {
            "ultima_fecha_procesada": ultima_fecha,
            "ultima_fecha_iso": None,
            "ultima_actualizacion": datetime.now().isoformat(),
        }

        if ultima_fecha:
            try:
                fecha_dt = datetime.strptime(ultima_fecha, "%d/%m/%y %H:%M:%S")
                estado["ultima_fecha_iso"] = fecha_dt.strftime("%Y-%m-%d")
            except ValueError:
                try:
                    fecha_dt = datetime.strptime(ultima_fecha.split()[0], "%d/%m/%y")
                    estado["ultima_fecha_iso"] = fecha_dt.strftime("%Y-%m-%d")
                except ValueError:
                    pass

        # Guardar estado en Supabase si es posible
        try:
            if self.supabase_manager:
                self.supabase_manager.guardar_estado_procesamiento(
                    estado["ultima_fecha_iso"]
                )
        except Exception as e:
            print(f"Error guardando estado en Supabase: {e}")

        try:
            os.makedirs("state", exist_ok=True)
            with open("state/last_processed.json", "w", encoding="utf-8") as f:
                json.dump(estado, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error guardando estado: {e}")

    def _obtener_fecha_guardada(self) -> Optional[str]:
        """Obtiene la fecha guardada del último procesamiento en formato YYYY-MM-DD."""
        # Priorizar estado persistido en Supabase
        try:
            if self.supabase_manager:
                fecha_supabase = self.supabase_manager.obtener_estado_procesamiento()
                if fecha_supabase:
                    return fecha_supabase
        except Exception as e:
            print(f"Error leyendo estado de Supabase: {e}")

        ruta_estado = os.path.join("state", "last_processed.json")

        if not os.path.exists(ruta_estado):
            return None

        try:
            with open(ruta_estado, "r", encoding="utf-8") as f:
                estado = json.load(f)

            fecha_iso = estado.get("ultima_fecha_iso")
            if fecha_iso:
                return fecha_iso

            # Intentar convertir el formato antiguo si existe
            fecha_original = estado.get("ultima_fecha_procesada")
            if not fecha_original:
                return None

            try:
                fecha_dt = datetime.strptime(fecha_original, "%d/%m/%y %H:%M:%S")
                return fecha_dt.strftime("%Y-%m-%d")
            except ValueError:
                try:
                    fecha_dt = datetime.strptime(fecha_original.split()[0], "%d/%m/%y")
                    return fecha_dt.strftime("%Y-%m-%d")
                except ValueError:
                    return None

        except (IOError, json.JSONDecodeError) as e:
            print(f"Error leyendo estado previo: {e}")
            return None


def main():
    """Función principal para ejecutar el extractor desde línea de comandos."""
    parser = argparse.ArgumentParser(
        description="Extraer recetas de archivos de WhatsApp"
    )
    parser.add_argument("--file", required=True, help="Ruta al archivo de WhatsApp")
    parser.add_argument(
        "--fecha-desde", help="Fecha desde la cual procesar (YYYY-MM-DD)"
    )

    args = parser.parse_args()

    # Cargar variables de entorno
    from dotenv import load_dotenv

    load_dotenv()

    # Crear extractor y procesar
    extractor = WhatsAppExtractor()

    fecha_desde = args.fecha_desde
    if not fecha_desde:
        fecha_desde = extractor._obtener_fecha_guardada()
        if fecha_desde:
            print(f"Usando fecha guardada por defecto: {fecha_desde}")

    resultado = extractor.procesar_archivo(args.file, fecha_desde)

    print("\n=== RESUMEN ===")
    print(f"Mensajes procesados: {resultado.get('mensajes_procesados', 0)}")
    print(f"Bloques procesados: {resultado.get('bloques_procesados', 0)}")
    print(f"Recetas extraídas: {resultado.get('recetas_extraidas', 0)}")
    print(f"Recetas insertadas: {resultado.get('recetas_insertadas', 0)}")


if __name__ == "__main__":
    main()
